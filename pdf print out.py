import tkinter as tk
from tkinter import messagebox, StringVar, ttk, Canvas, Frame
import bcrypt
import json
import os
import datetime
from PIL import Image, ImageTk
import openpyxl
import csv
import tkinter.ttk as ttk
from fpdf import FPDF

class LoginApp:
    def __init__(self, root):
        self.root = root
        self.root.title('Login')
        self.root.geometry('1166x718')
        self.root.config(bg='#001220')
        
        self.font1 = ('Helvetica', 25, 'bold')
        self.font2 = ('Arial', 17, 'bold')
        self.font3 = ('Arial', 13, 'bold')
        self.font4 = ('Arial', 17, 'bold', 'underline')
        self.font5 = ('Times New Roman', 14, 'bold')
        self.font6 = ('Baskerville', 14, 'bold')
        self.font7 = ('Sans Serif', 20, 'bold')
        self.font8 = ('Display', 15, 'bold')
        self.font9 = ('Display', 12, 'bold')

        self.current_user = None  # To store the logged-in username

        # Load the image and save reference
        self.image_path = os.path.join(os.path.dirname(__file__), "images", "Untitled design.png")
        self.photo = None
        self.load_image()

        self.show_signup_frame()

    def get_data_file_path(self):
        base_dir = os.path.dirname(__file__)
        return os.path.join(base_dir, "data.json")

    def save_data(self, data):
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        data["Waktu Presensi"] = current_time
        with open(self.get_data_file_path(), "w+") as file:
            json.dump(data, file, indent=4)

        self.save_data_to_excel()
        self.save_data_to_csv()

    def load_data(self):
        try:
            data_file_path = self.get_data_file_path()
            if not os.path.exists(data_file_path) or os.stat(data_file_path).st_size == 0:
                with open(data_file_path, "a") as file:
                    json.dump({}, file)
                return {}
            with open(data_file_path, "r") as file:
                return json.load(file)
        except (FileNotFoundError, json.JSONDecodeError):
            return {}
    
    def save_data_to_csv(self):
        data = self.load_data()
        filename = "data.csv"

        # Baca data yang sudah ada dari file CSV
        existing_data = []
        try:
            with open(filename, 'r', newline='') as file:
                reader = csv.DictReader(file)
                existing_data = list(reader)
        except FileNotFoundError:
            pass

        # Tambahkan data terbaru ke existing_data
        new_data = []
        for username, user_data in data.items():
            if isinstance(user_data, dict):
                row = {
                    'Username': username,
                    'Name': user_data.get('name', ''),
                    'NIM': user_data.get('nim', ''),
                    'Courses': ', '.join(user_data.get('courses', [])),
                    'Start Date': user_data.get('start_date', ''),
                    'Week Number': user_data.get('week_num', ''),
                    'Attendance Status': user_data.get('attendance', {}).get('attendance_status', ''),
                    'Reason': user_data.get('attendance', {}).get('reason', ''),
                    'Attendance Time': data.get('Waktu Presensi', ''),
                    'Summary': user_data.get('summary', '')
                }
                new_data.append(row)

        # Tulis data terbaru ke file CSV
        with open(filename, mode='w', newline='') as file:
            fieldnames = ['Username', 'Name', 'NIM', 'Courses', 'Start Date', 'Week Number', 'Attendance Status', 'Reason', 'Attendance Time', 'Summary']
            writer = csv.DictWriter(file, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(existing_data + new_data)

    def save_data_to_excel(self):
        data = self.load_data()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Tulis header
        worksheet['A1'] = 'Username'
        worksheet['B1'] = 'Name'
        worksheet['C1'] = 'NIM'
        worksheet['D1'] = 'Courses'
        worksheet['E1'] = 'Start Date'
        worksheet['F1'] = 'Week Number'
        worksheet['G1'] = 'Attendance Status'
        worksheet['H1'] = 'Reason'
        worksheet['I1'] = 'Summary'  # Tambahkan header untuk kolom 'Summary'

        row = 2
        for username, user_data in data.items():
            if isinstance(user_data, dict):
                worksheet.cell(row=row, column=1, value=username)
                worksheet.cell(row=row, column=2, value=user_data.get('name', ''))
                worksheet.cell(row=row, column=3, value=user_data.get('nim', ''))
                worksheet.cell(row=row, column=4, value=', '.join(user_data.get('courses', [])))
                worksheet.cell(row=row, column=5, value=user_data.get('start_date', ''))
                worksheet.cell(row=row, column=6, value=user_data.get('week_num', ''))
                attendance_data = user_data.get('attendance', {})
                worksheet.cell(row=row, column=7, value=attendance_data.get('attendance_status', ''))
                worksheet.cell(row=row, column=8, value=attendance_data.get('reason', ''))
                worksheet.cell(row=row, column=9, value=user_data.get('summary', ''))  # Simpan data rangkuman
            else:
                worksheet.cell(row=row, column=1, value=username)
            worksheet.cell(row=row, column=10, value=data.get('Waktu Presensi', ''))
            row += 1

        workbook.save('data.xlsx')
    def print_attendance_to_pdf(self):
        data = self.load_data()
        pdf = FPDF()

        new_data = []  # Daftar untuk menyimpan data presensi terbaru
        for username, user_data in data.items():
            if isinstance(user_data, dict):
                selected_courses = user_data.get('courses', [])
                attendance_time = data.get('Waktu Presensi', '')
                if attendance_time:
                    new_data.append((username, user_data, attendance_time, selected_courses))

        if new_data:
            for username, user_data, attendance_time, courses in new_data:
                pdf.add_page()
                pdf.set_font('Arial', 'B', 16)
                pdf.cell(0, 10, f'Presensi untuk {username}', ln=True)
                pdf.set_font('Arial', '', 12)
                pdf.cell(0, 8, f"Nama: {user_data.get('name', '')}", ln=True)
                pdf.cell(0, 8, f"NIM: {user_data.get('nim', '')}", ln=True)
                pdf.cell(0, 8, f"Mata Kuliah: {', '.join(selected_courses)}", ln=True)  # Hanya tampilkan mata kuliah yang dipilih
                pdf.cell(0, 8, f"Tanggal Mulai: {user_data.get('start_date', '')}", ln=True)
                pdf.cell(0, 8, f"Minggu Ke-: {user_data.get('week_num', '')}", ln=True)
                attendance_data = user_data.get('attendance', {})
                pdf.cell(0, 8, f"Status Kehadiran: {attendance_data.get('attendance_status', '')}", ln=True)
                pdf.cell(0, 8, f"Alasan: {attendance_data.get('reason', '')}", ln=True)
                pdf.cell(0, 8, f"Waktu Presensi: {attendance_time}", ln=True)
                pdf.cell(0, 8, f"Rangkuman: {user_data.get('summary', '')}", ln=True)

            pdf_file_path = 'presensi_terbaru.pdf'
            pdf.output(pdf_file_path, 'F')
            messagebox.showinfo('Success', 'File PDF presensi terbaru berhasil dibuat.')

            if messagebox.askyesno("Download PDF", "Apakah Anda ingin mendownload file PDF?"):
                import webbrowser
                webbrowser.open(pdf_file_path)

            self.show_login_frame()
        else:
            messagebox.showwarning('Warning', 'Tidak ada data presensi terbaru.')


        self.show_login_frame()
   
    def load_image(self):
        image = Image.open(self.image_path)
        image = image.resize((500, 718))  # Sesuaikan ukuran gambar
        self.photo = ImageTk.PhotoImage(image)

    def add_image_to_frame(self, frame):
        image_label = tk.Label(frame, image=self.photo, bg='#001220')
        image_label.place(relx=0.8, rely=0.5, anchor=tk.CENTER)

    def signup(self):
        username = self.username_entry.get()
        password = self.password_entry.get()
        data = self.load_data()
        if username != '' and password != '':
            if username in data:
                messagebox.showerror('Error', 'Username Telah Digunakan. Coba Username Lain.')
            else:
                hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
                data[username] = {"password": hashed_password.decode('utf-8'), "courses": []}
                self.save_data(data)
                messagebox.showinfo('Success', 'Akun Berhasil Dibuat')
                self.show_login_frame()
        else:
            messagebox.showerror('Error', 'Enter all data.')

    def save_courses(self):
        name = self.name_entry.get()
        nim = self.nim_entry.get()
        courses = [entry.get() for entry in self.course_entries if entry.get()]  # Hanya mengambil entri yang tidak kosong

    # Memeriksa apakah ada entri duplikat
        unique_courses = set(courses)
        if len(unique_courses) != len(courses):
            messagebox.showerror('Error', 'Tidak boleh ada course yang sama.')
            return

        data = self.load_data()
        if name and nim and self.current_user:
            if self.current_user in data:
                data[self.current_user]["name"] = name
                data[self.current_user]["nim"] = nim
                data[self.current_user]["courses"] = list(unique_courses)  # Simpan hanya mata kuliah yang unik dan dipilih
                self.save_data(data)
                messagebox.showinfo('Success', 'Courses and details have been saved.')
                self.show_course_selection_frame()  # Show the course selection frame after saving
            else:
                messagebox.showerror('Error', 'Invalid username.')
        else:
            messagebox.showerror('Error', 'Please fill in name and NIM.')
            
    def save_start_date_and_week_num(self):
        start_date = self.start_date_entry.get()
        week_num = self.week_num_entry.get()
        data = self.load_data()
        if start_date and week_num and self.current_user:
            if self.current_user in data:
                data[self.current_user]["start_date"] = start_date
                data[self.current_user]["week_num"] = week_num
                self.save_data(data)
                messagebox.showinfo('Success', 'Start date and week number have been saved.')
                self.show_attendance_frame()
            else:
                messagebox.showerror('Error', 'Invalid username.')
        else:
            messagebox.showerror('Error', 'Please fill in all fields.')

    def login(self):
        username = self.username_login_entry.get()
        password = self.password_login_entry.get()
        data = self.load_data()
        if username != '' and password != '':
            if username in data and bcrypt.checkpw(password.encode('utf-8'), data[username]['password'].encode('utf-8')):
                messagebox.showinfo('Success', 'Login Berhasil.')
                self.current_user = username  # Store the logged-in username
                self.show_course_input_frame()
                self.populate_course_data(data[username])  # Pre-populate course data
            else:
                messagebox.showerror('Error', 'Password atau Username Invalid')
        else:
            messagebox.showerror('Error', 'Enter all data.')

    def populate_course_data(self, user_data):
        self.name_entry.delete(0, 'end')
        self.name_entry.insert(0, user_data.get("name", ""))

        self.nim_entry.delete(0, 'end')
        self.nim_entry.insert(0, user_data.get("nim", ""))

        courses = user_data.get("courses", [])
        for i, entry in enumerate(self.course_entries):
            entry.delete(0, 'end')
            if i < len(courses):
                entry.insert(0, courses[i])

    def show_signup_frame(self):
        self.clear_frame()
        frame1 = tk.Frame(self.root, bg='#001220', width=583, height=718)
        frame1.place(x=0, y=0)

        signup_label = tk.Label(frame1, font=self.font6, text='Sign Up Akun Mahasiswa', bg='#001220', fg='#fff')
        signup_label.place(relx=0.5, rely=0.1, anchor=tk.CENTER)

        username_label = tk.Label(frame1, font=self.font6, text='Username:', bg='#001220', fg='#fff')
        username_label.place(relx=0.1, rely=0.3, anchor=tk.W)

        self.username_entry = tk.Entry(frame1, font=self.font5, fg='#FFFFFF', bg='#121111', insertbackground='#fff')
        self.username_entry.place(relx=0.3, rely=0.3, anchor=tk.W)

        password_label = tk.Label(frame1, font=self.font6, text='Password:', bg='#001220', fg='#fff')
        password_label.place(relx=0.1, rely=0.45, anchor=tk.W)

        self.password_entry = tk.Entry(frame1, font=self.font5, show='*', fg='#FFFFFF', bg='#121111', insertbackground='#fff')
        self.password_entry.place(relx=0.3, rely=0.45, anchor=tk.W)

        signup_button = tk.Button(frame1, command=self.signup, font=self.font6, text='Sign up', fg='#fff', bg='#00965d', activebackground='#006e44', width=12)
        signup_button.place(relx=0.5, rely=0.55, anchor=tk.CENTER)

        login_label = tk.Label(frame1, font=self.font6, text='Apakah sudah memiliki akun?', bg='#001220', fg='#fff')
        login_label.place(relx=0.5, rely=0.70, anchor=tk.CENTER)

        login_button = tk.Button(frame1, command=self.show_login_frame, font=self.font6, text='Login', fg='#fff', bg='#00965d', activebackground='#001220', width=8)
        login_button.place(relx=0.5, rely=0.75, anchor=tk.CENTER)
        
        self.add_image_to_frame(self.root)  # Menampilkan gambar di frame signup

    def show_print_pdf_frame(self):
        self.clear_frame()
        frame = tk.Frame(self.root, bg='#001220', width=700, height=600)
        frame.place(x=233, y=59)

        print_pdf_label = tk.Label(frame, font=self.font6, text='Print to PDF', bg='#001220', fg='#fff')
        print_pdf_label.place(relx=0.5, rely=0.1, anchor=tk.CENTER)

        print_pdf_button = tk.Button(frame, command=self.print_attendance_to_pdf, font=self.font8, text='Print PDF', fg='#fff', bg='#00965d', activebackground='#006e44', width=12)
        print_pdf_button.place(relx=0.5, rely=0.3, anchor=tk.CENTER)

        self.add_image_to_frame(self.root)

    def show_login_frame(self):
        self.clear_frame()
        frame2 = tk.Frame(self.root, bg='#001220', width=583, height=718)
        frame2.place(x=0, y=0)

        login_label2 = tk.Label(frame2, font=self.font6, text='Log in', bg='#001220', fg='#fff')
        login_label2.place(relx=0.5, rely=0.1, anchor=tk.CENTER)

        username_label2 = tk.Label(frame2, font=self.font6, text='Username:', bg='#001220', fg='#fff')  # Perbaikan penulisan
        username_label2.place(relx=0.1, rely=0.3, anchor=tk.W)

        password_label2 = tk.Label(frame2, font=self.font6, text='Password:', bg='#001220', fg='#fff')
        password_label2.place(relx=0.1, rely=0.45, anchor=tk.W)

        self.username_login_entry = tk.Entry(frame2, font=self.font5, fg='#fff', bg='#121111', insertbackground='#fff')
        self.username_login_entry.place(relx=0.475, rely=0.3, anchor=tk.CENTER)

        self.password_login_entry = tk.Entry(frame2, font=self.font5, show='*', fg='#fff', bg='#121111', insertbackground='#fff')
        self.password_login_entry.place(relx=0.475, rely=0.45, anchor=tk.CENTER)

        login_button2 = tk.Button(frame2, command=self.login, font=self.font6, text='Login', fg='#fff', bg='#00965d', activebackground='#006e44', width=12)
        login_button2.place(relx=0.5, rely=0.6, anchor=tk.CENTER)

        self.add_image_to_frame(self.root)  # Menampilkan gambar di frame login

    def show_course_input_frame(self):
        self.clear_frame()
        frame3 = tk.Frame(self.root, bg='#001220', width=700, height=600)
        frame3.place(x=233, y=59)

        name_label = tk.Label(frame3, font=self.font6, text='Name:', bg='#001220', fg='#fff')
        name_label.place(x=20, y=20)
        self.name_entry = tk.Entry(frame3, font=self.font5, fg='#FFFFFF', bg='#121111', insertbackground='#fff', width=30)
        self.name_entry.place(x=120, y=20)

        nim_label = tk.Label(frame3, font=self.font6, text='NIM:', bg='#001220', fg='#fff')
        nim_label.place(x=20, y=70)
        self.nim_entry = tk.Entry(frame3, font=self.font5, fg='#FFFFFF', bg='#121111', insertbackground='#fff', width=30)
        self.nim_entry.place(x=120, y=70)

        course_labels = ['Mata Kuliah 1:', 'Mata Kuliah 2:', 'Mata Kuliah 3:', 'Mata Kuliah 4:', 'Mata Kuliah 5:', 'Mata Kuliah 6:', 'Mata Kuliah 7:', 'Mata Kuliah 8:', 'Mata Kuliah 9:', 'Mata Kuliah 10:', 'Mata Kuliah 11:']
        self.course_entries = []

        for i, course_label in enumerate(course_labels):
            label = tk.Label(frame3, font=self.font9, text=course_label, bg='#001220', fg='#fff')
            label.place(x=20, y=120 + i * 30)
            entry = tk.Entry(frame3, font=self.font5, fg='#FFFFFF', bg='#121111', insertbackground='#fff', width=30)
            entry.place(x=120, y=120 + i * 30)
            self.course_entries.append(entry)

        self.save_button = tk.Button(frame3, command=self.save_courses, font=self.font8, text='Save', fg='#fff', bg='#00965d', activebackground='#006e44', width=12)
        self.save_button.place(relx=0.5, rely=0.9, anchor=tk.CENTER)
        
        self.add_image_to_frame(self.root)  # Menampilkan gambar di frame input course
        
    
    def show_course_selection_frame(self):
        self.clear_frame()
        frame4 = tk.Frame(self.root, bg='#ffffff', width=700, height=600)
        frame4.place(x=233, y=59)

        selection_label = tk.Label(frame4, font=self.font8, text='Pilih Mata Kuliah', bg='#ffffff', fg='#000000')
        selection_label.place(relx=0.3, rely=0.1, anchor=tk.CENTER)

        canvas = Canvas(frame4, bg='#ffffff', highlightthickness=0)
        canvas.place(relwidth=0.95, relheight=0.7, rely=0.2, relx=0.025)

        scrollbar = ttk.Scrollbar(frame4, orient="vertical", command=canvas.yview)
        scrollbar.place(relx=0.975, rely=0.2, relheight=0.7, anchor="ne")
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollable_frame = Frame(canvas, bg='#ffffff')
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        data = self.load_data()
        user_data = data.get(self.current_user, {})
        courses = user_data.get("courses", [])

        self.selected_courses = []

        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TRadiobutton', padx=20, pady=5, background='#ffffff', foreground='#000000', focuscolor='#3498db')
        style.map('TRadiobutton', background=[('active', '#3498db'), ('selected', '#3498db')], indicatorcolor=[('selected', '#ffffff'), ('active', '#ffffff')], indicatorrelief=[('selected', 'sunken')])

        for course in courses:
            var = StringVar()
            radiobutton = ttk.Radiobutton(scrollable_frame, text=course, variable=var, value=course, style='TRadiobutton', command=lambda: var.set(""))
            radiobutton.pack(anchor=tk.W, pady=5, padx=10)
            self.selected_courses.append(var)

        save_selection_button = tk.Button(frame4, command=self.show_start_date_and_week_num_frame, font=self.font8, text='Next', fg='#000000', bg='#ffffff', activebackground='#006e44', width=15)
        save_selection_button.place(relx=0.33, rely=0.95, anchor=tk.CENTER)

        self.add_image_to_frame(self.root)
    def save_selected_courses(self):
        selected_courses = [course.get() for course in self.selected_courses if course.get()]
        data = self.load_data()
        if self.current_user in data:
            data[self.current_user]["selected_courses"] = selected_courses
            self.save_data(data)
            messagebox.showinfo('Success', 'Selected courses have been saved.')
            self.show_start_date_and_week_num_frame()
        else:
            messagebox.showerror('Error', 'Invalid username.')   
    def show_start_date_and_week_num_frame(self):
        selected_courses = [var.get() for var in self.selected_courses if var.get()]
        data = self.load_data()

        if self.current_user in data:
            data[self.current_user]["selected_courses"] = selected_courses
            self.save_data(data)
        else:
            messagebox.showerror('Error', 'Invalid username.')

        self.clear_frame()
        frame5 = tk.Frame(self.root, bg='#001220', width=700, height=600)
        frame5.place(x=233, y=59)

        start_date_label = tk.Label(frame5, font=self.font6, text='Tanggal/waktu :', bg='#001220', fg='#fff')
        start_date_label.place(x=2, y=20)
        self.start_date_entry = tk.Entry(frame5, font=self.font6, fg='#FFFFFF', bg='#121111', insertbackground='#fff', width=30)
        self.start_date_entry.place(x=155, y=20)

        # Mengisi entri start date dengan waktu realtime
        current_datetime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.start_date_entry.insert(0, current_datetime)

        week_num_label = tk.Label(frame5, font=self.font6, text='Minggu ke- :', bg='#001220', fg='#fff')
        week_num_label.place(x=2, y=70)
        self.week_num_entry = tk.Entry(frame5, font=self.font5, fg='#FFFFFF', bg='#121111', insertbackground='#fff', width=30)
        self.week_num_entry.place(x=155, y=70)

        save_button = tk.Button(frame5, command=self.save_start_date_and_week_num, font=self.font8, text='Save', fg='#fff', bg='#00965d', activebackground='#006e44', width=12)
        save_button.place(relx=0.5, rely=0.9, anchor=tk.CENTER)

        self.add_image_to_frame(self.root)  # Menampilkan gambar di frame input start date dan week number

    def show_attendance_frame(self):
        self.clear_frame()
        frame6 = tk.Frame(self.root, bg='#001220', width=700, height=600)
        frame6.place(x=133, y=59)

        attendance_label = tk.Label(frame6, font=self.font6, text='Kehadiran', bg='#001220', fg='#fff')
        attendance_label.place(relx=0.5, rely=0.1, anchor=tk.CENTER)

        self.attendance_status = StringVar()  # Hapus nilai default 'Hadir'
        hadir_radio = ttk.Radiobutton(frame6, text="Hadir", value="Hadir", variable=self.attendance_status, style="TRadiobutton", command=self.save_attendance)
        hadir_radio.place(relx=0.4, rely=0.3, anchor=tk.CENTER)
        tidak_hadir_radio = ttk.Radiobutton(frame6, text="Tidak Hadir", value="Tidak Hadir", variable=self.attendance_status, style="TRadiobutton", command=self.toggle_reason_entry)
        tidak_hadir_radio.place(relx=0.6, rely=0.3, anchor=tk.CENTER)

        self.add_image_to_frame(self.root)  # Menampilkan gambar di frame input start date dan week number

    def show_summary_frame(self):
        self.clear_frame()
        frame8 = tk.Frame(self.root, bg='#001220', width=700, height=600)
        frame8.place(x=133, y=59)

        summary_label = tk.Label(frame8, font=self.font6, text='Input Rangkuman', bg='#001220', fg='#fff')
        summary_label.place(relx=0.5, rely=0.1, anchor=tk.CENTER)

        self.summary_text = tk.Text(frame8, font=self.font5, fg='#FFFFFF', bg='#121111', insertbackground='#fff', width=35, height=5)
        self.summary_text.place(relx=0.5, rely=0.4, anchor=tk.CENTER)

        save_summary_button = tk.Button(frame8, command=self.save_summary, font=self.font8, text='Save', fg='#fff', bg='#00965d', activebackground='#006e44', width=12)
        save_summary_button.place(relx=0.5, rely=0.7, anchor=tk.CENTER)

        print_pdf_button = tk.Button(frame8, command=self.show_print_pdf_frame, font=self.font8, text='Print PDF', fg='#fff', bg='#00965d', activebackground='#006e44', width=12)
        print_pdf_button.place(relx=0.5, rely=0.8, anchor=tk.CENTER)

        self.add_image_to_frame(self.root)

    def save_summary(self):
        summary_text = self.summary_text.get("1.0", "end-1c")  # Mendapatkan teks dari TextBox
        data = self.load_data()
        if self.current_user in data:
            data[self.current_user]["summary"] = summary_text
            self.save_data(data)
            messagebox.showinfo('Success', 'Rangkuman berhasil disimpan.')
        else:
            messagebox.showerror('Error', 'Invalid username.')
    
    def toggle_reason_entry(self):
        if self.attendance_status.get() == "Tidak Hadir":
            self.show_reason_frame()  # Tampilkan frame alasan ketidakhadiran
        else:
            self.reason_entry.grid_remove()  # Sembunyikan entri alasan ketidakhadiran

    def show_reason_frame(self):
        self.clear_frame()
        frame7 = tk.Frame(self.root, bg='#001220', width=700, height=600)
        frame7.place(x=133, y=59)

        reason_label = tk.Label(frame7, font=self.font6, text='Reason for Absence', bg='#001220', fg='#fff')
        reason_label.place(relx=0.5, rely=0.1, anchor=tk.CENTER)

        alasan_label = tk.Label(frame7, font=self.font6, text='Alasan :', bg='#001220', fg='#fff')
        alasan_label.place(relx=0.3, rely=0.4, anchor=tk.CENTER)

        reasons = ["Sakit", "Izin"]
        self.reason_var = tk.StringVar(self.root)
        self.reason_var.set(reasons[0])
        reason_dropdown = tk.OptionMenu(frame7, self.reason_var, *reasons)
        reason_dropdown.config(font=self.font5, fg='#FFFFFF', bg='#121111', activebackground='#001220')
        reason_dropdown.place(relx=0.5, rely=0.4, anchor=tk.CENTER)

        save_reason_button = tk.Button(frame7, command=self.save_attendance_with_reason, font=self.font8, text='Save', fg='#fff', bg='#00965d', activebackground='#006e44', width=12)
        save_reason_button.place(relx=0.5, rely=0.7, anchor=tk.CENTER)

        print_pdf_button = tk.Button(frame7, command=self.show_print_pdf_frame, font=self.font8, text='Print PDF', fg='#fff', bg='#00965d', activebackground='#006e44', width=12)
        print_pdf_button.place(relx=0.5, rely=0.8, anchor=tk.CENTER)

        self.add_image_to_frame(self.root)

    def save_attendance(self):
        attendance_status = self.attendance_status.get()
        if attendance_status == "Hadir":
            self.save_attendance_data(attendance_status, "")
            self.show_summary_frame()  # Tampilkan frame rangkuman jika hadir
        else:
            self.show_reason_frame()

    def save_attendance_with_reason(self):
        attendance_status = "Tidak Hadir"
        reason = self.reason_var.get()  # Get selected reason from the dropdown
        if reason == "Lainnya":
            custom_reason = self.reason_entry.get("1.0", "end-1c")  # If 'Lainnya' is selected, get custom reason from the entry
            if custom_reason:
                reason = custom_reason
            else:
                messagebox.showerror('Error', 'Silakan masukkan alasan ketidakhadiran.')
                return
        self.save_attendance_data(attendance_status, reason)
        self.show_reason_frame()  # Tetap di halaman alasan ketidakhadiran setelah menyimpan alasan


    def save_attendance_data(self, attendance_status, reason):
        data = self.load_data()
        if self.current_user in data:
            attendance_data = {
                "attendance_status": attendance_status,
                "reason": reason
            }
            data[self.current_user]["attendance"] = attendance_data
            self.save_data(data)
            
            # Hanya tampilkan pesan jika status kehadiran adalah "Tidak Hadir"
            if attendance_status == "Tidak Hadir":
                messagebox.showinfo('Success', 'Presensi berhasil disimpan.')
                self.show_attendance_frame()  # Tampilkan kembali frame attendance setelah menyimpan alasan
        else:
            messagebox.showerror('Error', 'Invalid username.')
    
    
    def clear_frame(self):
        for widget in self.root.winfo_children():
            widget.destroy()

if __name__ == '__main__':
    root = tk.Tk()
    app = LoginApp(root)                                                                                                                          
    root.mainloop()
